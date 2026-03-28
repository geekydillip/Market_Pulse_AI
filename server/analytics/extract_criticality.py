"""
extract_criticality.py
======================
Unified extraction for two file schemas.
Auto-detects which schema a file belongs to, then scores each issue
across its dimensions including the new 'similar_bug_count' parameter.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SCHEMA DETECTION
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    'Priority' + 'Occurr. Freq.' in columns  →  Schema 1
    'Resolve'  in columns (no Priority)      →  Schema 2

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SCHEMA 1  (e.g. high_issues.xlsx)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    Columns used  : Priority, Occurr. Freq., Issue Type, Progr.Stat.
    Dimensions    : priority (25%) + frequency (25%) + issue_type (20%)
                    + status (15%) + similar_bug_count (15%)

    Excluded rows (Progr.Stat.) :
        Resolve-Unnecessary, Resolve-Duplicated, Close, Closed

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SCHEMA 2  (e.g. high.xlsx, OneUI85.xlsx)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    Columns used  : Severity, Issue Type, Sub-Issue Type, Resolve
    Dimensions    : severity (35%) + issue_type (25%) + sub_issue (15%)
                    + resolve (10%) + similar_bug_count (15%)

    ── CHANGE: severity_filter removed — ALL severities are processed ──

    Excluded rows (Progr.Stat.) :
        Resolve-Unnecessary, Maintain current status, Not problem,
        Close, Closed

    Excluded rows (Resolve) :
        Duplicated issue(Cause side), Maintain current status, Not problem

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SIMILAR BUG COUNT  (new dimension, applied to both schemas)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    Two signals combined:

    1. combo_count — issues with identical (Issue Type + Sub-Issue Type).
    2. title_similar_count — TF-IDF cosine similarity on titles.

    similar_bug_count = max(combo_count, title_similar_count)
    Relative score = min(similar_bug_count, cap) / cap

    Falls back to combo_count only if sklearn is not installed.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SCORE  (0-100)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    score = sum(dimension_relative_score * weight) * 100
    Weights auto-normalise to 1.0.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
TIERS  (stable regardless of weight changes)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    Severe   >= 70
    Moderate >= 45
    Low       < 45
"""

import pandas as pd
import json
import numpy as np
from typing import Optional

try:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
    SKLEARN_AVAILABLE = True
except ImportError:
    SKLEARN_AVAILABLE = False


# ===========================================================================
#  SHARED CONFIG
# ===========================================================================

TIER_THRESHOLDS = {
    'Severe':   80,
    'Moderate': 45,
    # below 45 → Low
}

ISSUE_TYPE_RANK = {
    'System':        10,
    'Crash':         10,
    'App Crash':     10,
    'Security':       9,
    'Connectivity':   9,
    'Functional':     8,
    'Battery':        7,
    'Heat':           6,
    'Performance':    6,
    'Compatibility':  4,
    'UI/UX':          3,
    'Usability':      2,
    'UI':             2,
    'Other Issue':    1,
}

SUB_ISSUE_TYPE_RANK = {
    'App Crash':                  10,
    'CP Crash':                   10,
    'Power Off':                  10,
    'Restart':                     9,
    'No network':                  9,
    '5G Compatibility':            9,
    'Not Working':                 8,
    'Feature Missing':             7,
    'Battery Drain':               6,
    'Heating Issue':               6,
    'Weak Signal':                 5,
    'Compatibility Issue':         5,
    'Slow/Lag':                    4,
    'Slow/Lag Performance Issue':  4,
    'Performance':                 4,
    'Performance Issue':           4,
    'Poor Quality':                3,
    'UI Issue':                    2,
    'Display distortion':          2,
    'Other':                       1,
}

MAX_SIMILAR_BUGS = 20
TITLE_SIMILARITY_THRESHOLD = 0.25


# ===========================================================================
#  SCHEMA 1 CONFIG
# ===========================================================================

SCHEMA1_WEIGHTS = {
    'priority':   0.25,
    'frequency':  0.25,
    'issue_type': 0.20,
    'status':     0.15,
    'similar':    0.15,
}

SCHEMA1_PRIORITY_RANK = {
    'A': 3,
    'B': 2,
    'C': 1,
}

SCHEMA1_FREQUENCY_RANK = {
    'Always':    3,
    'Sometimes': 2,
    'Once':      1,
}

SCHEMA1_STATUS_RANK = {
    'Open':                    3,
    'Resolve - Not Released':  2,
    'Resolve - App Update':    1,
    'Resolve - Released':      1,
    'Close':                   0,
}

# ── UPDATED: added Close/Closed to Schema 1 exclusions ──────────────────────
SCHEMA1_EXCLUDED_STATUSES = {
    'Resolve - Unnecessary',
    'Resolve - Duplicated',
    'Close',
    'Closed',
}


# ===========================================================================
#  SCHEMA 2 CONFIG
# ===========================================================================

SCHEMA2_WEIGHTS = {
    'severity':   0.35,
    'issue_type': 0.25,
    'sub_issue':  0.15,
    'resolve':    0.10,
    'similar':    0.15,
}

SCHEMA2_SEVERITY_RANK = {
    'High':   3,
    'Medium': 2,
    'Low':    1,
}

SCHEMA2_RESOLVE_RANK = {
    'Issue Fixed(Source changes)':             4,
    'Issue Fixed(Except Source changes)':      4,
    'App Update via App Store':                3,
    'Not reproduced':                          1,
    'Insufficient Defect Info.':               1,
    'Request to 3rd Party(Non Samsung Issue)': 1,
    'Duplicated issue(Cause side)':            1,
    'PROCESSING':3,
    'RESOLVED':3,
    'OPENED':4,
}

# ── UPDATED: added Close/Closed to Progr.Stat. exclusions ───────────────────
SCHEMA2_EXCLUDED_STATUSES = {
    'Resolve - Unnecessary',
    'Maintain current status',
    'Not problem',
    'Close',
    'Closed',
    'CLOSED'
}

# ── UPDATED: added Not problem + Maintain current status to Resolve exclusions
SCHEMA2_EXCLUDED_RESOLVE = {
    'Maintain current status',
    'Not problem',
    'Duplicated issue(Cause side)',
}


# ===========================================================================
#  INTERNAL HELPERS
# ===========================================================================

def _detect_schema(df: pd.DataFrame) -> str:
    cols = set(df.columns)
    if 'Priority' in cols and 'Occurr. Freq.' in cols:
        return 'schema1'
    elif 'Resolve' in cols:
        return 'schema2'
    else:
        raise ValueError(
            f"Cannot detect schema. Columns found: {sorted(cols)}\n"
            "Expected ['Priority','Occurr. Freq.'] for Schema 1, "
            "or ['Resolve'] for Schema 2."
        )


def _normalise_weights(w: dict) -> dict:
    total = sum(w.values())
    if total == 0:
        raise ValueError("All weights are zero.")
    return {k: v / total for k, v in w.items()}


def _rel(value, rank_map: dict, default_rank: int = 1) -> float:
    rank = rank_map.get(value, default_rank)
    max_rank = max(rank_map.values()) if rank_map else 1
    return rank / max_rank


def _assign_tier(score: float, thresholds: dict) -> str:
    if score >= thresholds['Severe']:
        return 'Severe'
    elif score >= thresholds['Moderate']:
        return 'Moderate'
    return 'Low'


def _normalize_status(val):
    """Normalize Progr.Stat. into Open / Resolve / Close"""
    val = str(val).lower()
    if 'open' in val:
        return 'Open'
    elif 'close' in val:
        return 'Close'
    elif 'resolve' in val:
        return 'Resolve'
    else:
        return 'Other'


def _build_severity_summary(df: pd.DataFrame, excluded_df: pd.DataFrame = None) -> dict:
    """
    Flat summary keyed by severity level: High / Medium / Low.
    Each entry: { total, open, resolve, close }.

    Scale-down
    ----------
    HIGH   = High-severity rows that scored Severe tier (active only).
    MEDIUM = original Medium active
           + High rows scored Moderate  (active)
           + High rows scored Deferred  (active + excluded High rows)
    LOW    = Low-severity active
           + excluded rows that are NOT High-severity

    HIGH + MEDIUM + LOW = total_input (every row counted exactly once).

    Output
    ------
    {
        "High":   { "total": 60,  "open": 45, "resolve": 15, "close": 0,
                    "moved_to_medium": { "moderate": 55, "deferred": 127 } },
        "Medium": { "total": 804, "open": ..., "resolve": ..., "close": ... },
        "Low":    { "total": 439, "open": ..., "resolve": ..., "close": ... },
    }
    """
    if excluded_df is None:
        excluded_df = pd.DataFrame()

    def _counts(rows) -> dict:
        if len(rows) == 0:
            return {'total': 0, 'open': 0, 'resolve': 0, 'close': 0}
        sg = rows['status_group'] if 'status_group' in rows.columns else pd.Series(dtype=str)
        return {
            'total':   len(rows),
            'open':    int((sg == 'Open').sum()),
            'resolve': int((sg == 'Resolve').sum()),
            'close':   int((sg == 'Close').sum()),
        }

    def _add(dest: dict, src: dict):
        for k in ('total', 'open', 'resolve', 'close'):
            dest[k] = dest.get(k, 0) + src.get(k, 0)

    # ── Severity columns ──────────────────────────────────────────────────
    sev_col = (
        df['Severity'].fillna('Unknown').astype(str)
        if 'Severity' in df.columns
        else pd.Series(['Unknown'] * len(df), index=df.index)
    )
    excl_sev_col = (
        excluded_df['Severity'].fillna('Unknown').astype(str)
        if 'Severity' in excluded_df.columns and not excluded_df.empty
        else pd.Series(dtype=str)
    )

    # ── HIGH: Severe-tier active High rows only ───────────────────────────
    high_active   = df[sev_col == 'High']
    high_excluded = excluded_df[excl_sev_col == 'High'] if not excluded_df.empty else pd.DataFrame()

    high_severe   = high_active[high_active['tier'] == 'Severe']
    high_moderate = high_active[high_active['tier'] == 'Moderate']
    high_low_act  = high_active[high_active['tier'] == 'Low']

    # Deferred = active Low-tier High + excluded High
    high_deferred = pd.concat([high_low_act, high_excluded], ignore_index=True)

    high_entry = _counts(high_severe)
    high_entry['moved_to_medium'] = {
        'moderate': len(high_moderate),
        'deferred': len(high_deferred),
    }

    # ── MEDIUM: original medium active + moved from High ─────────────────
    medium_active = df[sev_col == 'Medium']
    medium_entry  = _counts(medium_active)
    _add(medium_entry, _counts(high_moderate))
    _add(medium_entry, _counts(high_deferred))

    # ── LOW: low active + excluded non-High rows ──────────────────────────
    low_active = df[sev_col == 'Low']
    non_high_excluded = excluded_df[excl_sev_col != 'High'] if not excluded_df.empty else pd.DataFrame()

    low_entry = _counts(low_active)
    _add(low_entry, _counts(non_high_excluded))

    return {
        'High':   high_entry,
        'Medium': medium_entry,
        'Low':    low_entry,
    }


# ===========================================================================
#  SIMILAR BUG COUNT
# ===========================================================================

def _compute_similar_bug_counts(
    df: pd.DataFrame,
    similarity_threshold: float = TITLE_SIMILARITY_THRESHOLD,
    max_similar_bugs: int = MAX_SIMILAR_BUGS,
) -> pd.Series:
    n = len(df)
    idx = df.index

    # Signal 1: combo count (Issue Type + Sub-Issue Type)
    combo_col = df['Issue Type'].astype(str) + '||' + df['Sub-Issue Type'].astype(str)
    combo_map = combo_col.value_counts().to_dict()
    combo_counts = combo_col.map(combo_map).fillna(1).astype(int)
    combo_counts = (combo_counts - 1).clip(lower=0)

    # Signal 2: TF-IDF title similarity
    if SKLEARN_AVAILABLE and n > 1:
        titles = df['Title'].fillna('').astype(str).tolist()
        try:
            tfidf = TfidfVectorizer(stop_words='english', ngram_range=(1, 2), min_df=1)
            mat = tfidf.fit_transform(titles)
            sim_matrix = cosine_similarity(mat)
            np.fill_diagonal(sim_matrix, 0)
            title_sim_counts = pd.Series(
                (sim_matrix >= similarity_threshold).sum(axis=1),
                index=idx, dtype=int
            )
        except Exception:
            title_sim_counts = pd.Series(0, index=idx, dtype=int)
    else:
        title_sim_counts = pd.Series(0, index=idx, dtype=int)

    combined = combo_counts.values + title_sim_counts.values
    capped = np.clip(combined, 0, max_similar_bugs)
    relative = capped / max_similar_bugs

    return pd.Series(relative, index=idx, name='_similar_rel')


# ===========================================================================
#  SCHEMA-SPECIFIC SCORERS
# ===========================================================================

def _score_schema1(df: pd.DataFrame, cfg: dict) -> pd.DataFrame:
    w   = cfg['weights']
    pr  = cfg['priority_rank']
    fr  = cfg['frequency_rank']
    itr = cfg['issue_type_rank']
    sr  = cfg['status_rank']
    sim_cfg = cfg['similar_bug_config']

    df['_p']   = df['Priority'].apply(lambda v: _rel(v, pr))
    df['_f']   = df['Occurr. Freq.'].apply(lambda v: _rel(v, fr))
    df['_i']   = df['Issue Type'].apply(lambda v: _rel(v, itr))
    df['_s']   = df['Progr.Stat.'].apply(lambda v: _rel(v, sr, default_rank=0))
    df['_sim'] = _compute_similar_bug_counts(
        df,
        similarity_threshold=sim_cfg['title_similarity_threshold'],
        max_similar_bugs=sim_cfg['max_similar_bugs'],
    )

    df['criticality_score'] = (
        df['_p']   * w['priority']   +
        df['_f']   * w['frequency']  +
        df['_i']   * w['issue_type'] +
        df['_s']   * w['status']     +
        df['_sim'] * w['similar']
    ) * 100

    df['priority_contribution']   = (df['_p']   * w['priority']   * 100).round(1)
    df['frequency_contribution']  = (df['_f']   * w['frequency']  * 100).round(1)
    df['issue_type_contribution'] = (df['_i']   * w['issue_type'] * 100).round(1)
    df['status_contribution']     = (df['_s']   * w['status']     * 100).round(1)
    df['similar_contribution']    = (df['_sim'] * w['similar']    * 100).round(1)

    df.drop(columns=['_p', '_f', '_i', '_s', '_sim'], inplace=True)
    return df


def _score_schema2(df: pd.DataFrame, cfg: dict) -> pd.DataFrame:
    w   = cfg['weights']
    svr = cfg['severity_rank']
    itr = cfg['issue_type_rank']
    sir = cfg['sub_issue_rank']
    rr  = cfg['resolve_rank']
    sim_cfg = cfg['similar_bug_config']

    df['_sv']  = df['Severity'].apply(lambda v: _rel(v, svr))
    df['_i']   = df['Issue Type'].apply(lambda v: _rel(v, itr))
    df['_si']  = df['Sub-Issue Type'].apply(lambda v: _rel(v, sir))
    df['_r']   = df['Resolve'].fillna('').apply(lambda v: _rel(v, rr, default_rank=1))
    df['_sim'] = _compute_similar_bug_counts(
        df,
        similarity_threshold=sim_cfg['title_similarity_threshold'],
        max_similar_bugs=sim_cfg['max_similar_bugs'],
    )

    df['criticality_score'] = (
        df['_sv']  * w['severity']   +
        df['_i']   * w['issue_type'] +
        df['_si']  * w['sub_issue']  +
        df['_r']   * w['resolve']    +
        df['_sim'] * w['similar']
    ) * 100

    df['severity_contribution']   = (df['_sv']  * w['severity']   * 100).round(1)
    df['issue_type_contribution'] = (df['_i']   * w['issue_type'] * 100).round(1)
    df['sub_issue_contribution']  = (df['_si']  * w['sub_issue']  * 100).round(1)
    df['resolve_contribution']    = (df['_r']   * w['resolve']    * 100).round(1)
    df['similar_contribution']    = (df['_sim'] * w['similar']    * 100).round(1)

    df.drop(columns=['_sv', '_i', '_si', '_r', '_sim'], inplace=True)
    return df


# ===========================================================================
#  CORE EXTRACTION FUNCTION
# ===========================================================================

def extract_criticality_data(
    filepath: str,
    # ── CHANGE: severity_filter removed — all severities are now processed ──
    # Similar bug config overrides
    max_similar_bugs: Optional[int] = None,
    title_similarity_threshold: Optional[float] = None,
    # Schema 1 overrides
    schema1_weights: Optional[dict] = None,
    schema1_priority_rank: Optional[dict] = None,
    schema1_frequency_rank: Optional[dict] = None,
    schema1_issue_type_rank: Optional[dict] = None,
    schema1_status_rank: Optional[dict] = None,
    # Schema 2 overrides
    schema2_weights: Optional[dict] = None,
    schema2_severity_rank: Optional[dict] = None,
    schema2_issue_type_rank: Optional[dict] = None,
    schema2_sub_issue_rank: Optional[dict] = None,
    schema2_resolve_rank: Optional[dict] = None,
    # Shared override
    tier_thresholds: Optional[dict] = None,
) -> dict:
    """
    Auto-detects file schema, excludes closed/unnecessary/duplicate rows,
    scores all remaining issues (all severities), and returns structured result.

    ── KEY CHANGE FROM PREVIOUS VERSION ────────────────────────────────────
    severity_filter parameter removed.
    All severities (High / Medium / Low) are now scored together.

    Exclusion rules (applied before scoring):
        Schema 1 Progr.Stat. :  Resolve-Unnecessary, Resolve-Duplicated,
                                 Close, Closed
        Schema 2 Progr.Stat. :  Resolve-Unnecessary, Maintain current status,
                                 Not problem, Close, Closed
        Schema 2 Resolve     :  Duplicated issue(Cause side),
                                 Maintain current status, Not problem
    ────────────────────────────────────────────────────────────────────────

    Returns
    -------
    dict:
        schema          → 'schema1' or 'schema2'
        issues          → list of scored dicts, sorted by score desc
        severity_breakdown → High/Medium/Low counts with scale-down applied
        total_input     → total rows in file
        excluded        → rows removed before scoring
        active          → rows that were scored
        config_used     → weights, thresholds, similar config applied
    """

    thr = tier_thresholds or TIER_THRESHOLDS

    sim_cfg = {
        'max_similar_bugs':           max_similar_bugs or MAX_SIMILAR_BUGS,
        'title_similarity_threshold': title_similarity_threshold or TITLE_SIMILARITY_THRESHOLD,
    }

    # ── 1. Load and detect schema ──────────────────────────────────────────
    df = pd.read_excel(filepath)
    total_input = len(df)
    schema = _detect_schema(df)

    # ── 2. Exclude non-actionable rows, build cfg, score ──────────────────
    if schema == 'schema1':

        excl_mask = df['Progr.Stat.'].isin(SCHEMA1_EXCLUDED_STATUSES)
        excluded_df = df[excl_mask].copy()
        excluded_count = len(excluded_df)
        df = df[~excl_mask].copy()

        # Normalize status for both sets
        df['status_group'] = df['Progr.Stat.'].apply(_normalize_status)
        excluded_df['status_group'] = excluded_df['Progr.Stat.'].apply(_normalize_status)

        cfg = {
            'weights':          _normalise_weights(schema1_weights or SCHEMA1_WEIGHTS),
            'priority_rank':    schema1_priority_rank   or SCHEMA1_PRIORITY_RANK,
            'frequency_rank':   schema1_frequency_rank  or SCHEMA1_FREQUENCY_RANK,
            'issue_type_rank':  schema1_issue_type_rank or ISSUE_TYPE_RANK,
            'status_rank':      schema1_status_rank     or SCHEMA1_STATUS_RANK,
            'similar_bug_config': sim_cfg,
        }

        df = _score_schema1(df, cfg)

        score_cols = [
            'criticality_score', 'tier',
            'priority_contribution', 'frequency_contribution',
            'issue_type_contribution', 'status_contribution',
            'similar_contribution',
        ]
        base_cols = [
            'Case Code', 'Source', 'Title', 'Priority', 'Occurr. Freq.',
            'Issue Type', 'Sub-Issue Type', 'Module', 'Progr.Stat.',
            'Ai Summary', 'Severity',
        ]

    else:  # schema2

        excl_mask = (
            df['Progr.Stat.'].isin(SCHEMA2_EXCLUDED_STATUSES) |
            df['Resolve'].isin(SCHEMA2_EXCLUDED_RESOLVE)
        )
        excluded_df = df[excl_mask].copy()
        excluded_count = len(excluded_df)
        df = df[~excl_mask].copy()

        # Normalize status for both sets
        df['status_group'] = df['Progr.Stat.'].apply(_normalize_status)
        excluded_df['status_group'] = excluded_df['Progr.Stat.'].apply(_normalize_status)

        cfg = {
            'weights':          _normalise_weights(schema2_weights or SCHEMA2_WEIGHTS),
            'severity_rank':    schema2_severity_rank   or SCHEMA2_SEVERITY_RANK,
            'issue_type_rank':  schema2_issue_type_rank or ISSUE_TYPE_RANK,
            'sub_issue_rank':   schema2_sub_issue_rank  or SUB_ISSUE_TYPE_RANK,
            'resolve_rank':     schema2_resolve_rank    or SCHEMA2_RESOLVE_RANK,
            'similar_bug_config': sim_cfg,
        }

        df = _score_schema2(df, cfg)

        score_cols = [
            'criticality_score', 'tier',
            'severity_contribution', 'issue_type_contribution',
            'sub_issue_contribution', 'resolve_contribution',
            'similar_contribution',
        ]
        base_cols = [
            'Case Code', 'Model No.', 'Title', 'Severity', 'Issue Type',
            'Sub-Issue Type', 'Module', 'Resolve', 'Progr.Stat.', 'Ai Summary',
        ]

    active_count = len(df)

    # ── 3. Round, assign tier, sort ────────────────────────────────────────
    df['criticality_score'] = df['criticality_score'].round(1)
    df['tier'] = df['criticality_score'].apply(lambda s: _assign_tier(s, thr))
    df['Ai Summary'] = df['Ai Summary'].fillna('')
    df = df.sort_values('criticality_score', ascending=False).reset_index(drop=True)

    # ── 4. Build output ────────────────────────────────────────────────────
    existing_base = [c for c in base_cols if c in df.columns]
    issues = df[existing_base + score_cols].to_dict(orient='records')

    # ── 5. Build per-severity breakdown (with High scale-down applied) ────
    severity_breakdown = _build_severity_summary(df, excluded_df)

    return {
        'schema':             schema,
        'issues':             issues,
        'severity_breakdown': severity_breakdown,
        'total_input':        total_input,
        'excluded':           excluded_count,
        'active':             active_count,
        'config_used': {
            'schema':             schema,
            'weights':            cfg['weights'],
            'tier_thresholds':    thr,
            'similar_bug_config': sim_cfg,
            'sklearn_available':  SKLEARN_AVAILABLE,
        },
    }


# ===========================================================================
#  HELPER FUNCTIONS
# ===========================================================================

def get_severe_issues(filepath: str, **kwargs) -> list:
    """Returns only Severe issues, sorted by score descending."""
    return [i for i in extract_criticality_data(filepath, **kwargs)['issues']
            if i['tier'] == 'Severe']


def get_issues_by_tier(filepath: str, tier: str, **kwargs) -> list:
    """Returns issues for a specific tier: 'Severe', 'Moderate', or 'Low'."""
    if tier not in ('Severe', 'Moderate', 'Low'):
        raise ValueError("tier must be 'Severe', 'Moderate', or 'Low'")
    return [i for i in extract_criticality_data(filepath, **kwargs)['issues']
            if i['tier'] == tier]


def get_summary(filepath: str, **kwargs) -> dict:
    """Returns severity breakdown with scale-down applied: High / Medium / Low."""
    return extract_criticality_data(filepath, **kwargs)['severity_breakdown']


def export_to_json(filepath: str, output_path: str, **kwargs) -> None:
    """
    Extracts data and saves full result to a JSON file.

    Also writes a companion *_severity_breakdown.json next to output_path
    containing only the per-severity criticality summary so it can be
    consumed directly by dashboards or reporting pipelines.
    """
    import os
    result = extract_criticality_data(filepath, **kwargs)

    # Full result
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=2, ensure_ascii=False)
    print(f"Saved (full)               -> {output_path}")

    # Severity breakdown only
    base, ext = os.path.splitext(output_path)
    breakdown_path = f"{base}_severity_breakdown{ext}"
    with open(breakdown_path, 'w', encoding='utf-8') as f:
        json.dump(result['severity_breakdown'], f, indent=2, ensure_ascii=False)
    print(f"Saved (severity breakdown) -> {breakdown_path}")


def export_to_excel(filepath: str, output_path: str, **kwargs) -> None:
    """Extracts data and saves scored issues to a new Excel file."""
    result = extract_criticality_data(filepath, **kwargs)
    pd.DataFrame(result['issues']).to_excel(output_path, index=False)
    print(f"Saved → {output_path}")


# ===========================================================================
#  MOVED ISSUES EXPORT — ALL rows with tier + updated_tier columns
# ===========================================================================

def export_moved_issues_updated_file(filepath: str, output_path: str = None, **kwargs) -> str:
    """
    Save ALL rows from the original Excel file to 'moved_issues_updated_file.xlsx'.

    Two new columns added:
        tier         = raw criticality tier scored by engine
                       (Severe / Moderate / Low / Excluded)

        updated_tier = tier after High-severity scale-down:
                       Severity=High, tier=Severe   → updated_tier=Severe   (kept)
                       Severity=High, tier=Moderate → updated_tier=Moderate (moved)
                       Severity=High, tier=Low      → updated_tier=Deferred (moved)
                       Severity=High, tier=Excluded → updated_tier=Deferred (moved)
                       Any other Severity row        → updated_tier=tier     (unchanged)

    A Remark column shows what happened:
        "Severe - No Change"
        "Moved to Moderate from High"
        "Moved to Deferred from High"
        "No Change"  (for all Medium / Low / Excluded non-High rows)

    Output: moved_issues_updated_file.xlsx  in same folder as input.

    Excel colour coding:
        Red    = Severity=High, updated_tier=Severe   (most critical, kept)
        Amber  = Severity=High, updated_tier=Moderate (moved down)
        Pink   = Severity=High, updated_tier=Deferred (moved down)
        White  = all other rows (Medium / Low)
    """
    import os

    # ── Output path ────────────────────────────────────────────────────
    if output_path is None:
        src_dir     = os.path.dirname(os.path.abspath(str(filepath)))
        output_path = os.path.join(src_dir, 'moved_issues_updated_file.xlsx')
    print(f"  [INFO] Saving to: {output_path}")

    thr     = kwargs.get('tier_thresholds') or TIER_THRESHOLDS
    sim_cfg = {
        'max_similar_bugs':           kwargs.get('max_similar_bugs')           or MAX_SIMILAR_BUGS,
        'title_similarity_threshold': kwargs.get('title_similarity_threshold') or TITLE_SIMILARITY_THRESHOLD,
    }

    # ── Load ALL rows, preserve original order ─────────────────────────
    df_all = pd.read_excel(filepath)
    df_all['_orig_idx'] = range(len(df_all))
    schema = _detect_schema(df_all.drop(columns=['_orig_idx']))

    # ── Split active vs excluded ───────────────────────────────────────
    if schema == 'schema1':
        excl_mask = df_all['Progr.Stat.'].isin(SCHEMA1_EXCLUDED_STATUSES)
        cfg = {
            'weights':          _normalise_weights(SCHEMA1_WEIGHTS),
            'priority_rank':    SCHEMA1_PRIORITY_RANK,
            'frequency_rank':   SCHEMA1_FREQUENCY_RANK,
            'issue_type_rank':  ISSUE_TYPE_RANK,
            'status_rank':      SCHEMA1_STATUS_RANK,
            'similar_bug_config': sim_cfg,
        }
    else:
        excl_mask = (
            df_all['Progr.Stat.'].isin(SCHEMA2_EXCLUDED_STATUSES) |
            df_all['Resolve'].isin(SCHEMA2_EXCLUDED_RESOLVE)
        )
        cfg = {
            'weights':          _normalise_weights(SCHEMA2_WEIGHTS),
            'severity_rank':    SCHEMA2_SEVERITY_RANK,
            'issue_type_rank':  ISSUE_TYPE_RANK,
            'sub_issue_rank':   SUB_ISSUE_TYPE_RANK,
            'resolve_rank':     SCHEMA2_RESOLVE_RANK,
            'similar_bug_config': sim_cfg,
        }

    active_df   = df_all[~excl_mask].copy()
    excluded_df = df_all[excl_mask].copy()

    # ── Score active rows ──────────────────────────────────────────────
    active_df['status_group'] = active_df['Progr.Stat.'].apply(_normalize_status)
    if schema == 'schema1':
        active_df = _score_schema1(active_df, cfg)
    else:
        active_df = _score_schema2(active_df, cfg)
    active_df['criticality_score'] = active_df['criticality_score'].round(1)
    active_df['tier'] = active_df['criticality_score'].apply(lambda s: _assign_tier(s, thr))

    # ── Mark excluded rows ─────────────────────────────────────────────
    excluded_df['tier']              = 'Excluded'
    excluded_df['criticality_score'] = None

    # ── Combine and restore original row order ─────────────────────────
    combined = pd.concat([active_df, excluded_df], ignore_index=True)
    combined.sort_values('_orig_idx', inplace=True)
    combined.reset_index(drop=True, inplace=True)

    has_sev = 'Severity' in combined.columns

    # ── Compute updated_tier ───────────────────────────────────────────
    # Rule (only applies to Severity=High rows):
    #   High + tier=Severe   → updated_tier=Severe   (stays as most critical)
    #   High + tier=Moderate → updated_tier=Moderate (moved down)
    #   High + tier=Low      → updated_tier=Deferred (moved down)
    #   High + tier=Excluded → updated_tier=Deferred (moved down)
    # All other rows: updated_tier = tier  (unchanged)
    def _get_updated_tier(row):
        sev  = str(row['Severity']).strip() if has_sev else ''
        tier = str(row['tier']).strip()
        if sev == 'High':
            if tier == 'Severe':
                return 'Severe'
            if tier == 'Moderate':
                return 'Moderate'
            if tier in ('Low', 'Excluded'):
                return 'Deferred'
        if sev == 'Medium' and tier == 'Severe':
            return 'Moderate'   # Medium severity cannot stay Severe → scaled to Moderate
        return tier   # all other rows unchanged

    combined['updated_tier'] = combined.apply(_get_updated_tier, axis=1)

    # ── Remark column ──────────────────────────────────────────────────
    def _get_remark(row):
        sev          = str(row['Severity']).strip() if has_sev else ''
        tier         = str(row['tier']).strip()
        updated_tier = str(row['updated_tier']).strip()
        if sev == 'High':
            if updated_tier == 'Severe':
                return 'Severe - No Change'
            if updated_tier == 'Moderate':
                return 'Moved to Moderate from High'
            if updated_tier == 'Deferred':
                return 'Moved to Deferred from High'
        if sev == 'Medium' and tier == 'Severe' and updated_tier == 'Moderate':
            return 'Moved to Moderate from Medium (scored Severe)'
        return 'No Change'

    combined['Remark'] = combined.apply(_get_remark, axis=1)

    # ── Drop internal helper columns ───────────────────────────────────
    drop_cols = ['_orig_idx', 'status_group',
                 '_p', '_f', '_i', '_s', '_sim', '_sv', '_si', '_r',
                 'priority_contribution', 'frequency_contribution',
                 'issue_type_contribution', 'status_contribution',
                 'severity_contribution', 'sub_issue_contribution',
                 'resolve_contribution', 'similar_contribution']
    combined.drop(columns=[c for c in drop_cols if c in combined.columns],
                  inplace=True)

    # ── Put new columns first: Remark | tier | updated_tier | score ────
    front = [c for c in ['Remark', 'tier', 'updated_tier', 'criticality_score']
             if c in combined.columns]
    rest  = [c for c in combined.columns if c not in front]
    combined = combined[front + rest]

    # ── Sort: High rows first (Severe → Moderate → Deferred), then rest ─
    remark_order = {
        'Severe - No Change':                          0,
        'Moved to Moderate from High':                 1,
        'Moved to Deferred from High':                 2,
        'Moved to Moderate from Medium (scored Severe)': 3,
        'No Change':                                   4,
    }
    combined['_sort'] = combined['Remark'].map(remark_order).fillna(3)
    combined.sort_values('_sort', inplace=True)
    combined.drop(columns=['_sort'], inplace=True)
    combined.reset_index(drop=True, inplace=True)

    # ── Write to Excel ─────────────────────────────────────────────────
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            combined.to_excel(writer, index=False, sheet_name='All Issues')
            ws = writer.sheets['All Issues']

            from openpyxl.styles import PatternFill, Font

            # Header row styling
            hdr_fill = PatternFill(fill_type='solid', fgColor='1E293B')
            hdr_font = Font(bold=True, color='FFFFFF')
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font

            # Row colours by Remark value
            colours = {
                'Severe - No Change':                          'FDECEA',   # light red   (High kept)
                'Moved to Moderate from High':                 'FFF3CD',   # light amber (moved)
                'Moved to Deferred from High':                 'FCE4EC',   # light pink  (moved)
                'Moved to Moderate from Medium (scored Severe)': 'E8F4FD', # light blue  (Medium scaled)
                'No Change':                                   'FFFFFF',   # white       (other rows)
            }
            remark_col_idx = combined.columns.get_loc('Remark') + 1  # 1-based

            for row_idx in range(2, len(combined) + 2):
                cell_val = ws.cell(row_idx, remark_col_idx).value or ''
                colour   = colours.get(cell_val, 'FFFFFF')
                fill     = PatternFill(fill_type='solid', fgColor=colour)
                for col_idx in range(1, len(combined.columns) + 1):
                    ws.cell(row_idx, col_idx).fill = fill

            # Auto column widths (max 48)
            for col_idx, col_name in enumerate(combined.columns, start=1):
                vals    = combined[col_name].astype(str)
                max_len = max(len(str(col_name)),
                              vals.str.len().max() if not vals.empty else 0)
                ws.column_dimensions[
                    ws.cell(1, col_idx).column_letter
                ].width = min(int(max_len) + 4, 48)

        print(f"  [OK] Saved successfully → {output_path}")

    except Exception as exc:
        import traceback
        print(f"  [ERROR] Could not save file: {exc}")
        print(f"  [ERROR] Target path: {output_path}")
        traceback.print_exc()
        raise

    # ── Console summary ────────────────────────────────────────────────
    rc = combined['Remark'].value_counts()
    print(f"\n{'='*60}")
    print(f"  moved_issues_updated_file.xlsx")
    print(f"{'='*60}")
    print(f"  Total rows : {len(combined)}")
    print(f"  ├─ Severe - No Change                    : {rc.get('Severe - No Change', 0)}")
    print(f"  ├─ Moved to Moderate from High           : {rc.get('Moved to Moderate from High', 0)}")
    print(f"  ├─ Moved to Deferred from High           : {rc.get('Moved to Deferred from High', 0)}")
    print(f"  ├─ Moved to Moderate from Medium(Severe) : {rc.get('Moved to Moderate from Medium (scored Severe)', 0)}")
    print(f"  └─ No Change                             : {rc.get('No Change', 0)}")
    print(f"{'='*60}\n")

    return output_path

# ===========================================================================
#  EXAMPLE USAGE
# ===========================================================================
if __name__ == '__main__':

    # files = ['D:\\saturday_backyp_MPAI\\Market_Pulse_AI_V1.1\\downloads\\beta_ut_voc\\Monika_Beta VOC_UT_Beta VOC_S25_20260223_121302_Processed.xlsx']
    files = ['D:\\saturday_backyp_MPAI\\test.xlsx']

    for filepath in files:
        print("=" * 72)
        print(f"File   : {filepath.split('/')[-1].split(chr(92))[-1]}")

        result = extract_criticality_data(filepath)

        print(f"Schema : {result['schema']}")
        print(f"Rows   : {result['total_input']} total  |  "
              f"{result['excluded']} excluded (Close/Unnecessary/Not problem)  |  "
              f"{result['active']} scored (all severities)")
        print(f"Weights: {result['config_used']['weights']}")
        print(f"Similar: cap={result['config_used']['similar_bug_config']['max_similar_bugs']}, "
              f"threshold={result['config_used']['similar_bug_config']['title_similarity_threshold']}, "
              f"sklearn={'yes' if result['config_used']['sklearn_available'] else 'no (combo only)'}")
        print()

        # ── Severity Breakdown (scale-down applied) ────────────────────────
        print("  -- Severity Breakdown (High scaled-down) ------------------------------")
        sb = result['severity_breakdown']
        for sev_label in ('High', 'Medium', 'Low'):
            entry = sb.get(sev_label, {})
            if not entry:
                continue
            moved = entry.get('moved_to_medium', {})
            note  = ''
            if moved:
                note = (f"  ← kept Severe only "
                        f"[moved Moderate:{moved['moderate']} "
                        f"+ Deferred:{moved['deferred']} → Medium]")
            print(f"  {sev_label:6s}  total={entry['total']:4d}  "
                  f"[open={entry['open']}, resolve={entry['resolve']}, close={entry['close']}]"
                  f"{note}")
        total_accounted = sum(sb.get(s, {}).get('total', 0) for s in ('High', 'Medium', 'Low'))
        print(f"  {'─'*62}")
        print(f"  Total accounted = {total_accounted}  |  input = {result['total_input']}  "
              f"{'✓' if total_accounted == result['total_input'] else '✗ MISMATCH'}")
        print()

        # ── Top 10 issues ──────────────────────────────────────────────────
        print(f"  {'Sev':<6} {'Title':<50} {'Score':>6}  {'Sim':>5}  Tier")
        print("  " + "─" * 76)
        for issue in result['issues'][:10]:
            sev = issue.get('Severity', '?')
            print(
                f"  {sev:<6} {issue['Title'][:49]:<50} "
                f"{issue['criticality_score']:>6.1f}  "
                f"{issue['similar_contribution']:>5.1f}  "
                f"{issue['tier']}"
            )
        print()

        # ── Export moved issues with tier vs updated_tier ─────────────────
        export_moved_issues_updated_file(filepath)