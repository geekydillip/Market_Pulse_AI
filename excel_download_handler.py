#!/usr/bin/env python3
"""
excel_download_handler.py

Python script to handle Excel file downloads for the Market Pulse AI dashboard.
Generates Excel files in different formats based on the requested data type.
"""

import pandas as pd
import json
import sys
import os
from datetime import datetime
from typing import Dict, List, Any, Optional
import argparse

def create_modules_summary_excel(data: List[Dict], filename: str = None) -> str:
    """
    Create Excel file in modules summary format matching the sample:
    Columns: model, module, count, rows, titleCount, modelNo, topIssueTitle
    """
    if not filename:
        now = datetime.now()
        date_str = now.strftime('%Y-%m-%d')
        filename = f'smplm_modules_summary_{date_str}.xlsx'

    # Transform data to match the expected format
    excel_data = []

    for item in data:
        # Extract model information
        model = item.get('model', item.get('modelNo', item.get('Model No.', '')))

        # Extract module information
        module = item.get('module', item.get('Module', ''))

        # Get count
        count = item.get('count', item.get('Count', 0))

        # For rows, we'll use the same as count for now
        rows = count

        # For titleCount, try to extract from topIssueTitle
        title_count = 0
        top_issue_title = item.get('topIssueTitle', item.get('Top Issue Title', ''))

        # Try to parse count from title if it contains count info
        if '|' in top_issue_title:
            # Handle format like "Issue title 5 | Another issue 3"
            parts = top_issue_title.split(' | ')
            title_count = len(parts)
        elif top_issue_title and top_issue_title != 'N/A':
            title_count = 1

        excel_data.append({
            'model': model,
            'module': module,
            'count': count,
            'rows': rows,
            'titleCount': title_count,
            'modelNo': model,
            'topIssueTitle': top_issue_title
        })

    # Create DataFrame and save to Excel
    df = pd.DataFrame(excel_data)

    # Create Excel writer with formatting
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Data', index=False)

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Data']

        # Remove gridlines
        worksheet.sheet_view.showGridLines = False

        # Set column widths
        column_widths = {
            'A': 22,  # model
            'B': 17,  # module
            'C': 9,  # count
            'D': 9,  # rows
            'E': 9,  # titleCount
            'F': 22,  # modelNo
            'G': 50   # topIssueTitle
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # Style headers
        from openpyxl.styles import Font, Alignment, PatternFill
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")  # Blue, Accent 1, 25% darker
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Style data rows
        data_font = Font(name="Calibri", size=11, color="000000")
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = data_font

                # Center align columns B, C, D, E, F (2, 3, 4, 5, 6)
                if col_num in [2, 3, 4, 5, 6]:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                # Add borders with light blue color
                from openpyxl.styles import Border, Side
                thin_border = Border(
                    left=Side(style='thin', color='ADD8E6'),
                    right=Side(style='thin', color='ADD8E6'),
                    top=Side(style='thin', color='ADD8E6'),
                    bottom=Side(style='thin', color='ADD8E6')
                )
                cell.border = thin_border

    return filename

def create_total_cases_excel(data: List[Dict], filename: str = None) -> str:
    """
    Create Excel file in total cases format matching the sample:
    Columns: Case Code, Model No., Progr.Stat., S/W Ver., Title, Problem,
             Resolve Option(Medium), Module, Sub-Module, Issue Type, Sub-Issue Type,
             Summarized Problem, Severity, Severity Reason, Resolve Type, R&D Comment
    """
    if not filename:
        now = datetime.now()
        date_str = now.strftime('%Y-%m-%d')
        filename = f'smplm_total_cases_{date_str}.xlsx'

    # Transform data to match the expected format
    excel_data = []

    for item in data:
        excel_data.append({
            'Case Code': item.get('Case Code', item.get('caseCode', '')),
            'Model No.': item.get('Model No.', item.get('modelNo', item.get('model', ''))),
            'Progr.Stat.': item.get('Progr.Stat.', item.get('progrStat', '')),
            'S/W Ver.': item.get('S/W Ver.', item.get('swVer', item.get('S/W Version', ''))),
            'Title': item.get('Title', item.get('title', '')),
            'Problem': item.get('Problem', item.get('problem', '')),
            'Resolve Option(Medium)': item.get('Resolve Option(Medium)', item.get('resolveOption', '')),
            'Module': item.get('Module', item.get('module', '')),
            'Sub-Module': item.get('Sub-Module', item.get('subModule', '')),
            'Issue Type': item.get('Issue Type', item.get('issueType', '')),
            'Sub-Issue Type': item.get('Sub-Issue Type', item.get('subIssueType', '')),
            'Summarized Problem': item.get('Summarized Problem', item.get('summarizedProblem', '')),
            'Severity': item.get('Severity', item.get('severity', '')),
            'Severity Reason': item.get('Severity Reason', item.get('severityReason', '')),
            'Resolve Type': item.get('Resolve Type', item.get('resolveType', '')),
            'R&D Comment': item.get('R&D Comment', item.get('rdComment', ''))
        })

    # Create DataFrame and save to Excel
    df = pd.DataFrame(excel_data)

    # Create Excel writer with formatting
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Data', index=False)

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Data']

        # Remove gridlines
        worksheet.sheet_view.showGridLines = False

        # Set column widths
        column_widths = {
            'A': 15,  # Case Code
            'B': 20,  # Model No.
            'C': 11,  # Progr.Stat.
            'D': 15,  # S/W Ver.
            'E': 40,  # Title
            'F': 40,  # Problem
            'G': 25,  # Resolve Option(Medium)
            'H': 20,  # Module
            'I': 20,  # Sub-Module
            'J': 15,  # Issue Type
            'K': 18,  # Sub-Issue Type
            'L': 40,  # Summarized Problem
            'M': 10,  # Severity
            'N': 20,  # Severity Reason
            'O': 15,  # Resolve Type
            'P': 40   # R&D Comment
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # Style headers
        from openpyxl.styles import Font, Alignment, PatternFill
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")  # Blue, Accent 1, 25% darker
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Style data rows
        data_font = Font(name="Calibri", size=11, color="000000")
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = data_font

                # Center align columns C, G, H, I, J, K, M, O (3, 7, 8, 9, 10, 11, 13, 15)
                if col_num in [3, 7, 8, 9, 10, 11, 13, 15]:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                # Add borders with light blue color
                from openpyxl.styles import Border, Side
                thin_border = Border(
                    left=Side(style='thin', color='ADD8E6'),
                    right=Side(style='thin', color='ADD8E6'),
                    top=Side(style='thin', color='ADD8E6'),
                    bottom=Side(style='thin', color='ADD8E6')
                )
                cell.border = thin_border

    return filename

def create_video_playback_excel(data: List[Dict], filename: str = None) -> str:
    """
    Create Excel file in video playback format (similar to total cases but focused on video issues)
    """
    # For now, use the same format as total cases since the structure is similar
    return create_total_cases_excel(data, filename)

def filter_high_severity(data: List[Dict]) -> List[Dict]:
    """Filter data to include only high severity items"""
    return [item for item in data if
            str(item.get('Severity', item.get('severity', ''))).lower() == 'high']

def main():
    parser = argparse.ArgumentParser(description='Generate Excel files for download')
    parser.add_argument('data_file', help='JSON file containing the data to export')
    parser.add_argument('export_type', choices=['modules', 'total', 'high', 'video'],
                       help='Type of export: modules (summary), total (all cases), high (high severity), video (video playback)')
    parser.add_argument('--output', '-o', help='Output filename (optional)')

    args = parser.parse_args()

    # Load data from JSON file
    try:
        with open(args.data_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        print(f"Error loading data file: {e}", file=sys.stderr)
        sys.exit(1)

    # Ensure data is a list
    if not isinstance(data, list):
        print("Data file must contain a JSON array", file=sys.stderr)
        sys.exit(1)

    # Filter data based on type
    if args.export_type == 'high':
        data = filter_high_severity(data)
        export_type = 'total'  # Use total format for high severity
    else:
        export_type = args.export_type

    # Generate Excel file
    try:
        if export_type == 'modules':
            filename = create_modules_summary_excel(data, args.output)
        elif export_type == 'total':
            filename = create_total_cases_excel(data, args.output)
        elif export_type == 'video':
            filename = create_video_playback_excel(data, args.output)

        print(f"Excel file created: {filename}")

    except Exception as e:
        print(f"Error creating Excel file: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    main()
